from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from pydantic import BaseModel
from typing import Dict, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os, threading, random, re, io, json, base64, queue
from datetime import datetime
from contextlib import asynccontextmanager
import requests as _req

BOT_TOKEN    = os.environ.get("TG_BOT_TOKEN", "8704747686:AAEL3E6QMKXvsznhxO5CKqumbSYHR3G4erM")
BACKEND_URL  = os.environ.get("BACKEND_URL", "https://chem-bio1.onrender.com")

_bot_application = None

async def _get_bot_app():
    global _bot_application
    if _bot_application is None:
        from telegram.ext import ApplicationBuilder, CommandHandler, CallbackQueryHandler, MessageHandler, filters
        import bot as _bot
        _bot_application = ApplicationBuilder().token(BOT_TOKEN).build()
        _bot_application.add_handler(CommandHandler("start", _bot.start))
        _bot_application.add_handler(CallbackQueryHandler(_bot.tugma))
        _bot_application.add_handler(MessageHandler(filters.PHOTO, _bot.rasm_handler))
        _bot_application.add_handler(MessageHandler(filters.Document.ALL, _bot.fayl_handler))
        _bot_application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, _bot.matn_handler))
        await _bot_application.initialize()
    return _bot_application

@asynccontextmanager
async def lifespan(application):
    _cache_load_all()
    try:
        _req.post(f"https://api.telegram.org/bot{BOT_TOKEN}/setWebhook",
                  json={"url": f"{BACKEND_URL}/webhook"}, timeout=10)
    except Exception:
        pass
    tw = threading.Thread(target=_gh_writer_loop, daemon=True)
    tw.start()
    tk = threading.Thread(target=_keep_alive_loop, daemon=True)
    tk.start()
    yield

app = FastAPI(lifespan=lifespan)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# ── GitHub storage ────────────────────────────────────────────────────────────
GH_TOKEN = os.environ.get("GH_TOKEN", "")
GH_REPO  = "asadullohnabiyev515-ux/chem-bio1"
GH_API   = "https://api.github.com"
GH_BRANCH = "db"

_cache: Dict[str, object] = {}
_cache_lock = threading.Lock()
_write_queue: queue.Queue = queue.Queue()

DATA_FILES = ["tests.json","students.json","roster.json","kitoblar.json",
              "balllar.json","savollar.json","elonlar.json","varaqlar.json","users.json"]

def _gh_headers():
    return {"Authorization": f"token {GH_TOKEN}", "Accept": "application/vnd.github.v3+json"}

def _gh_get(filename):
    try:
        r = _req.get(f"{GH_API}/repos/{GH_REPO}/contents/data/{filename}",
                     headers=_gh_headers(), params={"ref": GH_BRANCH}, timeout=15)
        if r.ok:
            raw = base64.b64decode(r.json()["content"]).decode()
            return json.loads(raw)
    except Exception:
        pass
    if filename in ("elonlar.json",): return []
    return {}

def _gh_put(filename, data):
    try:
        r = _req.get(f"{GH_API}/repos/{GH_REPO}/contents/data/{filename}",
                     headers=_gh_headers(), params={"ref": GH_BRANCH}, timeout=10)
        sha = r.json().get("sha") if r.ok else None
        content = base64.b64encode(json.dumps(data, ensure_ascii=False, indent=2).encode()).decode()
        payload = {"message": f"update {filename}", "content": content, "branch": GH_BRANCH}
        if sha: payload["sha"] = sha
        _req.put(f"{GH_API}/repos/{GH_REPO}/contents/data/{filename}",
                 headers=_gh_headers(), json=payload, timeout=20)
    except Exception as e:
        import logging; logging.getLogger(__name__).error(f"gh_put {filename}: {e}")

def _cache_load_all():
    for f in DATA_FILES:
        with _cache_lock:
            _cache[f] = _gh_get(f)

def _gh_writer_loop():
    while True:
        filename, data = _write_queue.get()
        _gh_put(filename, data)
        _write_queue.task_done()

def _keep_alive_loop():
    import time
    time.sleep(60)
    while True:
        try:
            _req.get(os.environ.get("BACKEND_URL", "https://chem-bio1.onrender.com") + "/", timeout=10)
        except Exception:
            pass
        time.sleep(300)

def load_json(filename):
    with _cache_lock:
        if filename not in _cache:
            _cache[filename] = _gh_get(filename)
        return _cache[filename]

def save_json(filename, data):
    with _cache_lock:
        _cache[filename] = data
    _write_queue.put((filename, data))

def ts():
    return datetime.now().strftime("%Y-%m-%d %H:%M")

# ── Ball yordamchi ────────────────────────────────────────────────────────────
def ball_qosh(user_id, ism, familiya, miqdor, sabab):
    balllar = load_json("balllar.json")
    uid = str(user_id)
    joriy_oy = datetime.now().strftime("%Y-%m")
    yozuv = {"miqdor": miqdor, "sabab": sabab, "vaqt": ts()}
    if uid not in balllar:
        balllar[uid] = {"ism": ism, "familiya": familiya, "jami": 0,
                        "oy_ball": 0, "oy": joriy_oy, "tarix": []}
    info = balllar[uid]
    if info.get("oy") != joriy_oy:
        info["oy_ball"] = 0; info["oy"] = joriy_oy
    info["jami"] = round(info.get("jami", 0) + miqdor, 1)
    info["oy_ball"] = round(info.get("oy_ball", 0) + miqdor, 1)
    info["ism"] = ism; info["familiya"] = familiya
    info.setdefault("tarix", []).append(yozuv)
    save_json("balllar.json", balllar)

# ── Models ────────────────────────────────────────────────────────────────────
class TestCreate(BaseModel):
    kod: str; fan: str; ustoz_id: int; javoblar: Dict[str, str]
    tur: str = "milliy"; savollar_soni: int = 40; vaqt: int = 90

class StudentRegister(BaseModel):
    test_kod: str; ism: str; familiya: str; telefon: str; user_id: int

class StudentResult(BaseModel):
    test_kod: str; ism: str; familiya: str; telefon: str
    user_id: int; javoblar: Dict[str, str]; varaq_id: Optional[str] = None

class FinishTest(BaseModel):
    test_kod: str; ustoz_id: int

class ScanResult(BaseModel):
    test_kod: str; varaq_id: str = ""; javoblar: Dict[str, str]
    ism: Optional[str] = None; familiya: Optional[str] = None

class KitobQosh(BaseModel):
    fan: str; nomi: str; fayl_id: str; ustoz_id: int

class KitobYuborish(BaseModel):
    fayl_id: str; user_id: int; nomi: str = "Kitob"

class SavolQosh(BaseModel):
    fan: str; mavzu: str; savol: str; variantlar: Dict[str, str]
    togri_javob: str; qiyinlik: str = "orta"; ustoz_id: int

class DosTTest(BaseModel):
    yaratuvchi_id: int; fan: str; savollar_soni: int = 10

# ── Asosiy ────────────────────────────────────────────────────────────────────
@app.get("/")
def root(): return {"status": "ok"}

@app.post("/webhook")
async def webhook(request: Request):
    from telegram import Update
    try:
        bot_app = await _get_bot_app()
        data = await request.json()
        update = Update.de_json(data, bot_app.bot)
        await bot_app.process_update(update)
    except Exception as e:
        import logging; logging.getLogger(__name__).error(f"Webhook xato: {e}")
        return {"ok": False, "error": str(e)}
    return {"ok": True}

@app.get("/debug/bot")
async def debug_bot():
    try:
        bot_app = await _get_bot_app()
        me = await bot_app.bot.get_me()
        wh = await bot_app.bot.get_webhook_info()
        return {"ok": True, "bot": me.username, "webhook_url": wh.url, "pending": wh.pending_update_count}
    except Exception as e:
        return {"ok": False, "error": str(e)}

# ── Testlar ───────────────────────────────────────────────────────────────────
@app.post("/test/yarat")
def test_yarat(data: TestCreate):
    tests = load_json("tests.json")
    asl_kod = data.kod; sessiya_num = 1; final_kod = asl_kod
    while final_kod in tests:
        sessiya_num += 1; final_kod = f"{asl_kod}-{sessiya_num}"
    tests[final_kod] = {
        "fan": data.fan, "ustoz_id": data.ustoz_id, "javoblar": data.javoblar,
        "yaratildi": ts(), "tugadi": False, "tur": data.tur,
        "savollar_soni": data.savollar_soni, "vaqt": data.vaqt,
        "asl_kod": asl_kod, "sessiya": sessiya_num
    }
    save_json("tests.json", tests)
    return {"success": True, "kod": final_kod, "sessiya": sessiya_num}

@app.get("/test/tekshir/{kod}")
def test_tekshir(kod: str):
    tests = load_json("tests.json")
    if kod not in tests: return {"topildi": False}
    t = tests[kod]
    return {"topildi": True, "fan": t["fan"], "tugadi": t["tugadi"],
            "tur": t.get("tur","milliy"), "savollar_soni": t.get("savollar_soni",40),
            "vaqt": t.get("vaqt",90), "javoblar": t.get("javoblar",{})}

@app.get("/ustoz/testlar/{ustoz_id}")
def ustoz_testlar(ustoz_id: int):
    tests = load_json("tests.json")
    students = load_json("students.json")
    roster = load_json("roster.json")
    natija = []
    for kod, t in tests.items():
        if t.get("ustoz_id") == ustoz_id:
            natija.append({
                "kod": kod, "fan": t["fan"], "yaratildi": t["yaratildi"],
                "tugadi": t["tugadi"], "talabalar": len(students.get(kod,[])),
                "royxat": len(roster.get(kod,{})),
                "tur": t.get("tur","milliy"), "savollar_soni": t.get("savollar_soni",40)
            })
    return {"testlar": natija}

@app.delete("/test/ochir/{kod}")
def test_ochir(kod: str, ustoz_id: int):
    tests = load_json("tests.json")
    if kod not in tests: raise HTTPException(404, "Test topilmadi")
    if tests[kod].get("ustoz_id") != ustoz_id: raise HTTPException(403, "Ruxsat yoq")
    del tests[kod]
    save_json("tests.json", tests)
    return {"success": True}

# ── O'quvchi ro'yxat ──────────────────────────────────────────────────────────
@app.post("/student/royxat")
def student_royxat(data: StudentRegister):
    tests = load_json("tests.json")
    if data.test_kod not in tests: raise HTTPException(404, "Test topilmadi")
    if tests[data.test_kod]["tugadi"]: raise HTTPException(400, "Test yakunlangan")
    roster = load_json("roster.json")
    if data.test_kod not in roster: roster[data.test_kod] = {}
    for vid, info in roster[data.test_kod].items():
        if info["user_id"] == data.user_id:
            return {"success": True, "varaq_id": vid, "xabar": "Allaqachon royxatdan otgansiz"}
    varaq_id = str(len(roster[data.test_kod]) + 1).zfill(3)
    roster[data.test_kod][varaq_id] = {
        "ism": data.ism, "familiya": data.familiya, "telefon": data.telefon,
        "user_id": data.user_id, "vaqt": ts()
    }
    save_json("roster.json", roster)
    return {"success": True, "varaq_id": varaq_id}

@app.get("/student/id/{test_kod}/{varaq_id}")
def student_id_get(test_kod: str, varaq_id: str):
    roster = load_json("roster.json")
    if test_kod not in roster: raise HTTPException(404, "Test topilmadi")
    vid = varaq_id.zfill(3)
    if vid not in roster[test_kod]: raise HTTPException(404, "Oquvchi topilmadi")
    return {"success": True, "oquvchi": roster[test_kod][vid]}

# ── Natijalar ─────────────────────────────────────────────────────────────────
@app.post("/natija/yuborish")
def natija_yubor(data: StudentResult):
    tests = load_json("tests.json")
    if data.test_kod not in tests: raise HTTPException(404, "Test topilmadi")
    test = tests[data.test_kod]
    students = load_json("students.json")
    if data.test_kod not in students: students[data.test_kod] = []
    birinchi = not any(s["user_id"] == data.user_id and data.user_id != 0
                       for s in students[data.test_kod])
    togri_javoblar = test.get("javoblar", {})
    savollar_soni = test.get("savollar_soni", 40)
    tur = test.get("tur", "milliy")
    togri_soni = 0; notogri = []; barcha_natija = []
    for i in range(1, savollar_soni + 1):
        k = str(i)
        oj = data.javoblar.get(k,"").strip().upper()
        tj = togri_javoblar.get(k,"").strip().upper()
        if tj and oj == tj: togri_soni += 1
        elif tj: notogri.append({"savol": i, "berilgan": oj or "?", "togri": tj})
        barcha_natija.append({"savol": i, "berilgan": oj or "?", "togri": tj, "xato": oj != tj})
    if tur == "milliy":
        ball = round((togri_soni / savollar_soni) * 75, 1) if savollar_soni else 0; maks_ball = 75
    else:
        ball = round((togri_soni / savollar_soni) * 100, 1) if savollar_soni else 0; maks_ball = 100
    students[data.test_kod].append({
        "ism": data.ism, "familiya": data.familiya, "telefon": data.telefon,
        "user_id": data.user_id, "javoblar": data.javoblar,
        "varaq_id": data.varaq_id or "", "vaqt": ts(),
        "togri_soni": togri_soni, "ball": ball
    })
    save_json("students.json", students)
    if birinchi and data.user_id and data.user_id != 0:
        ball_qosh(data.user_id, data.ism, data.familiya, ball,
                  f"Test: {data.test_kod} ({togri_soni}/{savollar_soni})")
    return {"success": True, "togri_soni": togri_soni, "jami": savollar_soni,
            "ball": ball, "maks_ball": maks_ball, "notogri": notogri,
            "barcha": barcha_natija,
            "foiz": round((togri_soni/savollar_soni)*100,1) if savollar_soni else 0,
            "birinchi_marta": birinchi}

@app.get("/student/qidir/{test_kod}")
def student_qidir(test_kod: str, q: str = ""):
    roster = load_json("roster.json")
    if test_kod not in roster or not q.strip(): return {"talabalar": []}
    q_lower = q.lower().strip()
    natija = []
    for vid, info in roster[test_kod].items():
        ism_full = f"{info.get('ism','')} {info.get('familiya','')}".lower().strip()
        if q_lower in ism_full:
            natija.append({"varaq_id": vid, "ism": info.get("ism",""),
                           "familiya": info.get("familiya",""),
                           "user_id": info.get("user_id",0)})
    return {"talabalar": natija[:10]}

@app.post("/natija/skaner")
def natija_skaner(data: ScanResult):
    tests = load_json("tests.json")
    if data.test_kod not in tests: raise HTTPException(404, "Test topilmadi")
    test = tests[data.test_kod]
    togri_javoblar = test.get("javoblar", {})
    n_savol = test.get("savollar_soni", 40)
    tur = test.get("tur", "milliy")
    roster = load_json("roster.json")
    students = load_json("students.json")
    if data.test_kod not in students: students[data.test_kod] = []
    oquvchi = None; vid = ""
    if data.varaq_id:
        vid = data.varaq_id.zfill(3)
        if data.test_kod in roster and vid in roster[data.test_kod]:
            oquvchi = roster[data.test_kod][vid]
    if not oquvchi:
        if not data.ism: raise HTTPException(400, "O'quvchi topilmadi")
        oquvchi = {"ism": data.ism, "familiya": data.familiya or "", "telefon": "", "user_id": 0}
    if vid:
        for s in students[data.test_kod]:
            if s.get("varaq_id") == vid:
                raise HTTPException(400, f"{oquvchi['ism']} allaqachon topshirgan!")
    togri_soni = sum(1 for i in range(1, n_savol + 1)
        if data.javoblar.get(str(i), "").upper() == togri_javoblar.get(str(i), "").upper())
    maks = 75 if tur == "milliy" else 100
    ball = round((togri_soni / n_savol) * maks, 1) if n_savol else 0
    students[data.test_kod].append({
        "ism": oquvchi["ism"], "familiya": oquvchi.get("familiya",""),
        "telefon": oquvchi.get("telefon",""), "user_id": oquvchi.get("user_id",0),
        "javoblar": data.javoblar, "varaq_id": vid, "vaqt": ts(),
        "togri_soni": togri_soni, "ball": ball
    })
    save_json("students.json", students)
    uid = oquvchi.get("user_id", 0)
    if uid and uid != 0:
        xabar = (f"📊 Skanerlash natijasi\n\n"
                 f"👤 {oquvchi['ism']} {oquvchi.get('familiya','')}\n"
                 f"📚 Fan: {test.get('fan','')}\n"
                 f"🔑 Test: {data.test_kod}\n"
                 f"✅ To'g'ri: {togri_soni}/{n_savol}\n"
                 f"🏆 Ball: {ball}/{maks}")
        try:
            _req.post(f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                json={"chat_id": uid, "text": xabar}, timeout=5)
        except Exception:
            pass
    return {"success": True, "oquvchi": oquvchi, "togri_soni": togri_soni, "ball": ball}

# ── Test yakunlash ────────────────────────────────────────────────────────────
@app.post("/test/yakunla")
def test_yakunla(data: FinishTest):
    tests = load_json("tests.json")
    students = load_json("students.json")
    if data.test_kod not in tests: raise HTTPException(404, "Test topilmadi")
    t = tests[data.test_kod]
    if t["ustoz_id"] != data.ustoz_id: raise HTTPException(403, "Ruxsat yoq")
    if t["tugadi"]: raise HTTPException(400, "Allaqachon yakunlangan")
    talabalar = students.get(data.test_kod, [])
    if not talabalar: raise HTTPException(400, "Natija yoq")
    togri_javoblar = t["javoblar"]; tur = t.get("tur","milliy"); n_savol = t.get("savollar_soni",40)
    n = len(talabalar)
    if tur == "milliy":
        togri_soni_dict = {}
        for i in range(1, n_savol+1):
            k = str(i)
            togri_soni_dict[k] = sum(1 for s in talabalar
                if s["javoblar"].get(k,"").strip().upper() == togri_javoblar.get(k,"").strip().upper())
        qiyinlik = {str(i): (1-(togri_soni_dict[str(i)]/n if n>0 else 0.5))+0.01
                    for i in range(1, n_savol+1)}
        maks_mumkin = sum(qiyinlik.values())
        def norm(xom): return round((xom/maks_mumkin)*75, 2) if maks_mumkin>0 else 0
        natijalar = []
        for s in talabalar:
            xom = sum(qiyinlik[str(i)] for i in range(1,n_savol+1)
                if s["javoblar"].get(str(i),"").strip().upper()==togri_javoblar.get(str(i),"").strip().upper())
            togri = sum(1 for j in range(1,n_savol+1)
                if s["javoblar"].get(str(j),"").strip().upper()==togri_javoblar.get(str(j),"").strip().upper())
            natijalar.append({**s, "togri_soni": togri, "ball": norm(xom)})
        maks_ball = 75
    else:
        natijalar = []
        for s in talabalar:
            togri = sum(1 for j in range(1,n_savol+1)
                if s["javoblar"].get(str(j),"").strip().upper()==togri_javoblar.get(str(j),"").strip().upper())
            natijalar.append({**s, "togri_soni": togri, "ball": round((togri/n_savol)*100,2)})
        maks_ball = 100
    natijalar.sort(key=lambda x: x["ball"], reverse=True)
    for idx, nn in enumerate(natijalar): nn["orin"] = idx+1
    excel_path = excel_yarat(data.test_kod, t["fan"], natijalar, togri_javoblar, tur, n_savol, maks_ball)
    tests[data.test_kod]["tugadi"] = True
    save_json("tests.json", tests)
    for nn in natijalar:
        uid = nn.get("user_id",0)
        if not uid or uid==0: continue
        orin = nn.get("orin",99)
        bonus = {1:30,2:20,3:10}.get(orin,0)
        ball_qosh(uid, nn.get("ism",""), nn.get("familiya",""), 5+bonus,
                  f"Test topshirish" + (f" ({orin}-orin +{bonus})" if bonus else ""))
    try:
        with open(excel_path, "rb") as ef:
            _req.post(f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument",
                files={"document": (excel_path, ef,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                data={"chat_id": data.ustoz_id,
                      "caption": f"📊 {t['fan']} — {data.test_kod}\n👥 {n} ta o'quvchi natijasi"},
                timeout=20)
    except Exception:
        pass
    return {"success": True, "excel_path": excel_path, "jami": n,
            "natijalar": natijalar, "togri_javoblar": togri_javoblar, "maks_ball": maks_ball}

# ── Excel ─────────────────────────────────────────────────────────────────────
def excel_yarat(kod, fan, natijalar, togri, tur, n_savol, maks_ball):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Natijalar"
    bold=Font(bold=True); center=Alignment(horizontal="center",vertical="center")
    SARIQ=PatternFill("solid",fgColor="FFF59D"); YASHIL=PatternFill("solid",fgColor="C8E6C9")
    QIZIL=PatternFill("solid",fgColor="FFCDD2"); KOK=PatternFill("solid",fgColor="BBDEFB")
    ws.merge_cells("A1:G1")
    ws["A1"] = ("Milliy Sertifikat (RASh)" if tur=="milliy" else f"Oddiy test ({n_savol} savol)") + f" — {fan} (Kod: {kod})"
    ws["A1"].font=Font(bold=True,size=13); ws["A1"].alignment=center; ws["A1"].fill=KOK
    for col,s in enumerate(["Orin","ID","Ism","Familiya","Telefon",f"Togri ({n_savol})",f"Ball ({maks_ball})"],1):
        c=ws.cell(row=2,column=col,value=s); c.font=bold; c.fill=SARIQ; c.alignment=center
    chegara=maks_ball*0.8; orta=maks_ball*0.5
    for ri,nn in enumerate(natijalar,3):
        fill=YASHIL if nn["ball"]>=chegara else (PatternFill("solid",fgColor="FFF9C4") if nn["ball"]>=orta else QIZIL)
        for ci,val in enumerate([nn["orin"],nn.get("varaq_id",""),nn["ism"],nn.get("familiya",""),
                                  nn.get("telefon",""),nn["togri_soni"],nn["ball"]],1):
            c=ws.cell(row=ri,column=ci,value=val); c.alignment=center; c.fill=fill
    for col in ["A","B","C","D","E","F","G"]: ws.column_dimensions[col].width=16
    path=f"natija_{kod}.xlsx"; wb.save(path); return path

@app.get("/excel/{kod}")
def excel_ol(kod: str):
    path=f"natija_{kod}.xlsx"
    if not os.path.exists(path): raise HTTPException(404,"Excel topilmadi")
    return FileResponse(path,filename=f"natija_{kod}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Kitoblar ──────────────────────────────────────────────────────────────────
@app.post("/kitob/qosh")
def kitob_qosh(data: KitobQosh):
    kitoblar=load_json("kitoblar.json")
    if data.fan not in kitoblar: kitoblar[data.fan]=[]
    kitoblar[data.fan].append({"nomi":data.nomi,"fayl_id":data.fayl_id,
                                "ustoz_id":data.ustoz_id,"vaqt":ts()})
    save_json("kitoblar.json",kitoblar); return {"success":True}

@app.get("/kitob/fanlar")
def kitob_fanlar():
    return {"fanlar": list(load_json("kitoblar.json").keys())}

@app.get("/kitob/royxat/{fan}")
def kitob_royxat(fan: str):
    kitoblar=load_json("kitoblar.json")
    return {"kitoblar": kitoblar.get(fan,[])}

@app.post("/kitob/yuborish")
def kitob_yuborish(data: KitobYuborish):
    try:
        r=_req.post(f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument",
            json={"chat_id":data.user_id,"document":data.fayl_id,"caption":f"📚 {data.nomi}"},timeout=10)
        if r.ok: return {"success":True}
        return {"success":False,"detail":r.json().get("description","Xato")}
    except Exception as e: raise HTTPException(500,str(e))

@app.post("/kitob/yukla-fayl")
async def kitob_yukla_fayl(
    fan: str = Form(...), nomi: str = Form(...), ustoz_id: int = Form(...),
    fayl: UploadFile = File(...)
):
    fayl_bytes = await fayl.read()
    r = _req.post(f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument",
        files={"document": (fayl.filename or "kitob.pdf", fayl_bytes)},
        data={"chat_id": ustoz_id}, timeout=30)
    if not r.ok: raise HTTPException(400, r.json().get("description", "Telegram xato"))
    file_id = r.json()["result"]["document"]["file_id"]
    kitoblar = load_json("kitoblar.json")
    if fan not in kitoblar: kitoblar[fan] = []
    kitoblar[fan].append({"nomi": nomi, "fayl_id": file_id, "ustoz_id": ustoz_id, "vaqt": ts()})
    save_json("kitoblar.json", kitoblar)
    return {"success": True, "fayl_id": file_id}

@app.post("/varaqlar/yukla-fayl")
async def varaqlar_yukla_fayl(
    soni: str = Form(...), ustoz_id: int = Form(...),
    fayl: UploadFile = File(...)
):
    fayl_bytes = await fayl.read()
    r = _req.post(f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument",
        files={"document": (fayl.filename or f"varaq_{soni}.pdf", fayl_bytes)},
        data={"chat_id": ustoz_id}, timeout=30)
    if not r.ok: raise HTTPException(400, r.json().get("description", "Telegram xato"))
    file_id = r.json()["result"]["document"]["file_id"]
    varaqlar = load_json("varaqlar.json")
    varaqlar[str(soni)] = {"fayl_id": file_id, "nomi": f"{soni} savollik varaq"}
    save_json("varaqlar.json", varaqlar)
    return {"success": True}

@app.delete("/kitob/ochir/{fan}/{idx}")
def kitob_ochir(fan: str, idx: int, ustoz_id: int):
    kitoblar=load_json("kitoblar.json")
    if fan not in kitoblar: raise HTTPException(404,"Fan topilmadi")
    if idx>=len(kitoblar[fan]): raise HTTPException(404,"Kitob topilmadi")
    kitoblar[fan].pop(idx)
    if not kitoblar[fan]: del kitoblar[fan]
    save_json("kitoblar.json",kitoblar); return {"success":True}

# ── Profil ────────────────────────────────────────────────────────────────────
@app.get("/profil/{user_id}")
def profil(user_id: int):
    students=load_json("students.json"); tests=load_json("tests.json"); natija=[]
    for test_kod,talabalar in students.items():
        for s in talabalar:
            if s.get("user_id")==user_id:
                t=tests.get(test_kod,{})
                natija.append({"test_kod":test_kod,"fan":t.get("fan",""),
                    "ball":s.get("ball",0),"togri_soni":s.get("togri_soni",0),
                    "vaqt":s.get("vaqt",""),"tugadi":t.get("tugadi",False)})
    natija.sort(key=lambda x: x["vaqt"],reverse=True)
    tugagan=[n for n in natija if n["tugadi"] and n.get("ball",0)>0]
    ortacha=round(sum(n["ball"] for n in tugagan)/len(tugagan),1) if tugagan else 0
    return {"jami_test":len(natija),"ortacha_ball":ortacha,
            "eng_yuqori":max((n["ball"] for n in tugagan),default=0),"natijalar":natija[:10]}

# ── Reyting ───────────────────────────────────────────────────────────────────
@app.get("/reyting/umumiy")
def reyting_umumiy():
    balllar=load_json("balllar.json"); natija=[]
    for uid,info in balllar.items():
        natija.append({"user_id":int(uid),"ism":info["ism"],"familiya":info.get("familiya",""),
                       "jami_ball":info["jami"]})
    natija.sort(key=lambda x: x["jami_ball"],reverse=True)
    for idx,n in enumerate(natija): n["orin"]=idx+1
    return {"reyting":natija[:20]}

@app.get("/reyting/oxirgi")
def reyting_oxirgi():
    tests=load_json("tests.json"); students=load_json("students.json")
    yakunlangan=[(k,v) for k,v in tests.items() if v.get("tugadi")]
    if not yakunlangan: return {"reyting":[],"test_kod":None,"fan":None}
    kod,test=sorted(yakunlangan,key=lambda x:x[1].get("yaratildi",""),reverse=True)[0]
    talabalar=students.get(kod,[])
    natija=[{"ism":s["ism"],"familiya":s.get("familiya",""),"ball":s["ball"],
             "togri_soni":s.get("togri_soni",0)} for s in talabalar if s.get("ball",0)>0]
    natija.sort(key=lambda x: x["ball"],reverse=True)
    for idx,n in enumerate(natija): n["orin"]=idx+1
    return {"reyting":natija[:20],"test_kod":kod,"fan":test.get("fan","")}

@app.get("/ball/tarix/{user_id}")
def ball_tarix(user_id: int):
    balllar=load_json("balllar.json"); uid=str(user_id)
    if uid not in balllar: return {"jami":0,"tarix":[]}
    info=balllar[uid]
    return {"jami":info.get("jami",0),"oy_ball":info.get("oy_ball",0),
            "tarix":list(reversed(info.get("tarix",[])))[:20]}

# ── Savollar bazasi ───────────────────────────────────────────────────────────
@app.post("/savol/qosh")
def savol_qosh(data: SavolQosh):
    savollar=load_json("savollar.json")
    if data.fan not in savollar: savollar[data.fan]={}
    if data.mavzu not in savollar[data.fan]: savollar[data.fan][data.mavzu]=[]
    savol_id=len(savollar[data.fan][data.mavzu])+1
    savollar[data.fan][data.mavzu].append({
        "id":savol_id,"savol":data.savol,"variantlar":data.variantlar,
        "togri_javob":data.togri_javob,"qiyinlik":data.qiyinlik,
        "ustoz_id":data.ustoz_id,"vaqt":ts()
    })
    save_json("savollar.json",savollar); return {"success":True,"savol_id":savol_id}

@app.get("/savol/fanlar")
def savol_fanlar():
    savollar=load_json("savollar.json"); natija=[]
    for fan,mavzular in savollar.items():
        jami=sum(len(v) for v in mavzular.values())
        natija.append({"fan":fan,"mavzular":list(mavzular.keys()),"jami":jami})
    return {"fanlar":natija}

@app.get("/savol/mavzular/{fan}")
def savol_mavzular(fan: str):
    savollar=load_json("savollar.json")
    if fan not in savollar: return {"mavzular":[]}
    return {"mavzular":[{"mavzu":m,"soni":len(s)} for m,s in savollar[fan].items()]}

@app.get("/savol/olish/{fan}/{mavzu}")
def savol_olish(fan: str, mavzu: str, soni: int=10):
    savollar=load_json("savollar.json")
    if fan not in savollar or mavzu not in savollar[fan]: return {"savollar":[]}
    all_s=savollar[fan][mavzu]
    tanlangan=random.sample(all_s,min(soni,len(all_s)))
    return {"savollar":[{"id":s["id"],"savol":s["savol"],"variantlar":s["variantlar"],
                         "qiyinlik":s["qiyinlik"]} for s in tanlangan]}

@app.post("/savol/tekshir")
def savol_tekshir(fan: str, mavzu: str, javoblar: Dict[str,str]):
    savollar=load_json("savollar.json")
    if fan not in savollar or mavzu not in savollar[fan]: raise HTTPException(404,"Savollar topilmadi")
    savol_map={str(s["id"]):s["togri_javob"] for s in savollar[fan][mavzu]}
    togri=0; notogri=[]
    for sid,javob in javoblar.items():
        tj=savol_map.get(sid,"")
        if javob.strip().upper()==tj.strip().upper(): togri+=1
        else: notogri.append({"savol_id":sid,"berilgan":javob,"togri":tj})
    ball=round((togri/len(javoblar))*100,1) if javoblar else 0
    return {"togri":togri,"jami":len(javoblar),"ball":ball,"notogri":notogri}

@app.post("/savol/pdf")
async def savol_pdf(
    fan: str = Form(...),
    mavzu: str = Form(...),
    ustoz_id: int = Form(0),
    fayl: UploadFile = File(...)
):
    import io
    fayl_bytes = await fayl.read()
    fname = (fayl.filename or "").lower()
    matn = ""
    if fname.endswith(".docx"):
        try:
            from docx import Document
            doc = Document(io.BytesIO(fayl_bytes))
            matn = "\n".join(p.text for p in doc.paragraphs)
        except Exception as e:
            raise HTTPException(400, f"Word fayl o'qib bo'lmadi: {e}")
    else:
        try:
            import fitz
            doc = fitz.open(stream=fayl_bytes, filetype="pdf")
            for page in doc:
                matn += page.get_text() + "\n"
            doc.close()
        except Exception:
            try:
                import pdfplumber
                with pdfplumber.open(io.BytesIO(fayl_bytes)) as pdf:
                    for page in pdf.pages:
                        text = page.extract_text()
                        if text:
                            matn += text + "\n"
            except Exception as e:
                raise HTTPException(400, f"PDF o'qib bo'lmadi: {e}")
    if not matn.strip():
        raise HTTPException(400, "Fayldan matn topilmadi")

    # Javoblar kalit — bir nechta formatni qo'llab-quvvatlaydi
    javoblar_kalit = {}

    # FORMAT A: "Javoblar: 1-A, 2-D, 3-B..."
    jav_m = re.search(r"[Jj]avoblar\s*:?\s*([\s\S]+?)(?:\n\n|\Z)", matn)
    if jav_m:
        for item in re.split(r"[,;\s]+", jav_m.group(1)):
            m2 = re.match(r"(\d+)\s*[-–.]\s*([A-Fa-f])", item.strip())
            if m2:
                javoblar_kalit[int(m2.group(1))] = m2.group(2).upper()

    # FORMAT B: Jadval — raqamlar qatori, keyin javoblar qatori
    # Masalan: "1  2  3  4  5\nd  D  B  C  B"
    if not javoblar_kalit:
        qatorlar = [q.strip() for q in matn.replace("\r","").split("\n") if q.strip()]
        for i in range(len(qatorlar) - 1):
            raqam_q = qatorlar[i]
            javob_q = qatorlar[i + 1]
            raqamlar = re.findall(r'\b(\d+)\b', raqam_q)
            javoblar = re.findall(r'\b([A-Fa-f])\b', javob_q)
            # Tekshirish: raqamlar ketma-ket va javoblar soni mos
            if (len(raqamlar) >= 3 and len(raqamlar) == len(javoblar)
                    and not re.search(r'[A-Za-z]{2,}', raqam_q)):
                for r, j in zip(raqamlar, javoblar):
                    javoblar_kalit[int(r)] = j.upper()

    savollar = load_json("savollar.json")
    if fan not in savollar: savollar[fan] = {}
    if mavzu not in savollar[fan]: savollar[fan][mavzu] = []
    qoshildi = 0

    def savol_qosh(num, savol_t, variantlar, togri):
        nonlocal qoshildi
        t = togri or javoblar_kalit.get(num, "")
        if savol_t and len(variantlar) >= 2 and t:
            sid = len(savollar[fan][mavzu]) + 1
            savollar[fan][mavzu].append({"id": sid, "savol": savol_t,
                "variantlar": variantlar, "togri_javob": t, "qiyinlik": "orta"})
            qoshildi += 1

    # --- FORMAT 1: Har bir savol alohida qatorda (klassik) ---
    # Savol: ...\nA) ...\nTogri: A
    joriy_savol = ""; joriy_variantlar = {}; joriy_togri = ""; joriy_num = 0
    klassik_topildi = False

    for qator in matn.replace("\r", "").split("\n"):
        q = qator.strip()
        if not q: continue
        sm  = re.match(r"^[Ss]avol\s*:\s*(.+)", q)
        nm  = re.match(r"^(\d+)\s*[.]\s+(.+)", q)
        vm  = re.match(r"^([A-D])\s*[)]\s*(.+)", q)
        tm  = re.match(r"^(?:[Tt]og.ri|[Jj]avob)\s*:\s*([A-D])", q)
        if sm:
            klassik_topildi = True
            if joriy_savol: savol_qosh(joriy_num, joriy_savol, dict(joriy_variantlar), joriy_togri)
            joriy_savol = sm.group(1).strip(); joriy_variantlar = {}; joriy_togri = ""
        elif nm:
            if joriy_savol: savol_qosh(joriy_num, joriy_savol, dict(joriy_variantlar), joriy_togri)
            joriy_num = int(nm.group(1))
            joriy_savol = nm.group(2).strip(); joriy_variantlar = {}; joriy_togri = ""
        elif vm:
            joriy_variantlar[vm.group(1)] = vm.group(2).strip()
        elif tm:
            joriy_togri = tm.group(1).upper()
    if joriy_savol: savol_qosh(joriy_num, joriy_savol, dict(joriy_variantlar), joriy_togri)

    # --- FORMAT 2: Bir qatorda "1. Savol matni A) ... B) ... C) ... D) ..." ---
    if qoshildi == 0:
        # Har bir raqamli blokni ajratib olamiz
        bloklar = re.split(r'(?:^|\n)(\d+)\.\s+', matn, flags=re.MULTILINE)
        # bloklar = ['prefix', '1', 'block1', '2', 'block2', ...]
        i = 1
        while i < len(bloklar) - 1:
            try:
                num = int(bloklar[i])
                blok = bloklar[i + 1]
                # A) B) C) D) bo'yicha ajratamiz (faqat katta harf)
                qismlar = re.split(r'\s+([A-D])\)\s+', blok)
                # qismlar = [savol_matni, 'A', a_text, 'B', b_text, 'C', c_text, 'D', d_text]
                if len(qismlar) >= 9:
                    savol_t = qismlar[0].strip()
                    variantlar = {}
                    for j in range(1, min(9, len(qismlar) - 1), 2):
                        variantlar[qismlar[j]] = qismlar[j + 1].strip()
                    togri = javoblar_kalit.get(num, "")
                    savol_qosh(num, savol_t, variantlar, togri)
            except Exception:
                pass
            i += 2

    save_json("savollar.json", savollar)
    # Agar 0 bo'lsa debug uchun birinchi 500 belgini qaytaramiz
    preview = matn[:500] if qoshildi == 0 else ""
    return {"success": True, "qoshildi": qoshildi, "matn_uzunlik": len(matn), "preview": preview}

# ── Varaq skaneri API ────────────────────────────────────────────────────────
def _varaq_skaner(rasm_bytes, n_savol=40):
    try:
        import numpy as np
        import cv2
    except ImportError:
        return None, None, "CV2 kutubxonasi yo'q"
    nparr = np.frombuffer(rasm_bytes, np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    if img is None: return None, None, "Rasm o'qilmadi"
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
    gray = clahe.apply(gray)
    blurred = cv2.GaussianBlur(gray, (5,5), 0)
    circles = cv2.HoughCircles(blurred, cv2.HOUGH_GRADIENT, dp=1.2, minDist=12,
        param1=50, param2=22, minRadius=7, maxRadius=18)
    if circles is None: return None, None, "Doiralar topilmadi. Rasmni aniqroq oling."
    circles = np.round(circles[0,:]).astype("int")
    def is_filled(g, cx, cy, r):
        mask = np.zeros(g.shape, dtype=np.uint8)
        cv2.circle(mask, (cx,cy), max(r-2,3), 255, -1)
        pixels = g[mask==255]
        return len(pixels) > 0 and np.sum(pixels<110)/len(pixels) > 0.33
    rows = {}
    for (cx,cy,r) in circles:
        k = cy // 15
        if k not in rows: rows[k] = []
        rows[k].append((cx,cy,r))
    sorted_rows = sorted(rows.items())
    varaq_id = ""
    id_digits = []
    for rk, row in sorted_rows[:4]:
        for (cx,cy,r) in sorted(row, key=lambda c: c[0]):
            if is_filled(gray,cx,cy,r): id_digits.append(str(rk%10))
    if id_digits: varaq_id = "".join(id_digits[:3]).zfill(3)
    results = {}; savol = 1
    for rk, row in sorted_rows:
        row_s = sorted(row, key=lambda c: c[0])
        if len(row_s) >= 4:
            harflar = ["A","B","C","D","E","F"] if 33<=savol<=35 else ["A","B","C","D"]
            for idx,(cx,cy,r) in enumerate(row_s[:len(harflar)]):
                if is_filled(gray,cx,cy,r): results[str(savol)] = harflar[idx]; break
            if str(savol) not in results: results[str(savol)] = ""
            savol += 1
            if savol > n_savol: break
    return results, varaq_id, None

@app.post("/skaner/rasm")
async def skaner_rasm_api(
    test_kod: str = Form(...),
    n_savol: int = Form(40),
    rasm: UploadFile = File(...)
):
    rasm_bytes = await rasm.read()
    javoblar, varaq_id, xato = _varaq_skaner(rasm_bytes, n_savol)
    if xato: raise HTTPException(400, xato)
    oquvchi = None
    if varaq_id:
        roster = load_json("roster.json")
        if test_kod in roster and varaq_id in roster[test_kod]:
            oquvchi = roster[test_kod][varaq_id]
    tests = load_json("tests.json")
    togri_javoblar = tests.get(test_kod, {}).get("javoblar", {})
    return {"success": True, "javoblar": javoblar, "varaq_id": varaq_id,
            "oquvchi": oquvchi, "togri_javoblar": togri_javoblar}


@app.post("/test/fayldan")
def test_fayldan(data: dict):
    fan=data.get("fan","").strip(); kod=data.get("kod","").strip().upper()
    matn=data.get("matn",""); ustoz_id=data.get("ustoz_id",0)
    vaqt=int(data.get("vaqt",90)); tur=data.get("tur","milliy")
    if not fan or not kod or not matn: raise HTTPException(400,"fan, kod va matn kerak")
    savollar=[]; javoblar={}
    joriy_savol=""; joriy_variantlar={}; joriy_togri=""
    def saqlash():
        nonlocal joriy_savol,joriy_variantlar,joriy_togri
        if joriy_savol and len(joriy_variantlar)>=2:
            n=len(savollar)+1
            savollar.append({"n":n,"savol":joriy_savol,"variantlar":dict(joriy_variantlar)})
            if joriy_togri: javoblar[str(n)]=joriy_togri
        joriy_savol=""; joriy_variantlar={}; joriy_togri=""
    for qator in matn.replace("\r","").split("\n"):
        qator=qator.strip()
        if not qator: continue
        savol_m=re.match(r"^[Ss]avol\s*:\s*(.+)",qator)
        num_m=re.match(r"^(\d+)\s*[.)\-]\s*(.+)",qator)
        var_m=re.match(r"^([A-Fa-f])\s*[).\-]\s*(.+)",qator)
        tog_m=re.match(r"^[Tt]og.?ri\s*:\s*([A-Fa-f])",qator)
        if savol_m:
            saqlash(); joriy_savol=savol_m.group(1).strip(); joriy_variantlar={}; joriy_togri=""
        elif num_m and not var_m:
            saqlash(); joriy_savol=re.sub(r"\s*\(\d+\)\s*$","",num_m.group(2)).strip(); joriy_variantlar={}; joriy_togri=""
        elif var_m: joriy_variantlar[var_m.group(1).upper()]=var_m.group(2).strip()
        elif tog_m: joriy_togri=tog_m.group(1).upper()
    saqlash()
    if not savollar: raise HTTPException(400,"Savollar topilmadi. Formatni tekshiring.")
    tests=load_json("tests.json")
    asl_kod=kod; sessiya=1; final_kod=asl_kod
    while final_kod in tests:
        sessiya+=1; final_kod=f"{asl_kod}-{sessiya}"
    tests[final_kod]={"fan":fan,"ustoz_id":ustoz_id,"javoblar":javoblar,
        "yaratildi":ts(),"tugadi":False,"tur":tur,"savollar_soni":len(savollar),
        "vaqt":vaqt,"asl_kod":asl_kod,"sessiya":sessiya,"savollar":savollar}
    save_json("tests.json",tests)
    return {"success":True,"kod":final_kod,"savollar_soni":len(savollar),"javoblar_soni":len(javoblar)}

@app.post("/test/savol/yukla")
def test_savol_yukla(data: dict):
    test_kod=data.get("test_kod","").strip(); matn=data.get("matn",""); ustoz_id=data.get("ustoz_id",0)
    if not test_kod or not matn: raise HTTPException(400,"test_kod va matn kerak")
    tests=load_json("tests.json")
    if test_kod not in tests: raise HTTPException(404,"Test topilmadi")
    if tests[test_kod].get("ustoz_id")!=ustoz_id: raise HTTPException(403,"Ruxsat yo'q")
    savollar=[]; joriy_savol=""; joriy_variantlar={}
    def saqlash():
        if joriy_savol and len(joriy_variantlar)>=2:
            savollar.append({"n":len(savollar)+1,"savol":joriy_savol,"variantlar":dict(joriy_variantlar)})
    for qator in matn.replace("\r","").split("\n"):
        qator=qator.strip()
        if not qator: continue
        savol_m=re.match(r"^[Ss]avol\s*:\s*(.+)",qator)
        num_m=re.match(r"^(\d+)\s*[.)\-]\s*(.+)",qator)
        var_m=re.match(r"^([A-Fa-f])\s*[).\-]\s*(.+)",qator)
        if savol_m:
            saqlash(); joriy_savol=savol_m.group(1).strip(); joriy_variantlar={}
        elif num_m and not var_m:
            saqlash(); joriy_savol=re.sub(r"\s*\(\d+\)\s*$","",num_m.group(2)).strip(); joriy_variantlar={}
        elif var_m:
            joriy_variantlar[var_m.group(1).upper()]=var_m.group(2).strip()
    saqlash()
    tests[test_kod]["savollar"]=savollar
    save_json("tests.json",tests)
    return {"success":True,"qoshildi":len(savollar)}

@app.get("/test/savollar/{kod}")
def test_savollar_get(kod: str):
    tests=load_json("tests.json")
    if kod not in tests: raise HTTPException(404,"Test topilmadi")
    savollar=tests[kod].get("savollar",[])
    return {"savollar":savollar,"bor":len(savollar)>0}

@app.delete("/savol/ochir/{fan}/{mavzu}")
def savol_ochir(fan: str, mavzu: str, ustoz_id: int):
    savollar=load_json("savollar.json")
    if fan not in savollar or mavzu not in savollar[fan]: raise HTTPException(404,"Topilmadi")
    del savollar[fan][mavzu]
    if not savollar[fan]: del savollar[fan]
    save_json("savollar.json",savollar); return {"success":True}

# ── E'lonlar ──────────────────────────────────────────────────────────────────
def _tg_xabar_yuborish(user_id: int, matn: str):
    try:
        _req.post(f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                  json={"chat_id": user_id, "text": matn, "parse_mode": "HTML"}, timeout=10)
    except Exception:
        pass

def _elon_bildirishnoma(elon: dict):
    users = load_json("users.json")
    if not isinstance(users, dict): return
    matn = (f"📢 <b>Yangi e'lon!</b>\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"<b>{elon['sarlavha']}</b>\n\n"
            f"{elon['matn']}\n\n"
            f"🕐 {elon['vaqt']}")
    for uid_str in users:
        try:
            _tg_xabar_yuborish(int(uid_str), matn)
        except Exception:
            pass

@app.post("/user/royxat")
def user_royxat(data: dict):
    users = load_json("users.json")
    if not isinstance(users, dict): users = {}
    uid = str(data.get("user_id", ""))
    if not uid: return {"success": False}
    users[uid] = {"ism": data.get("ism", ""), "vaqt": ts()}
    save_json("users.json", users)
    return {"success": True}

@app.post("/elon/qosh")
def elon_qosh(data: dict):
    elonlar=load_json("elonlar.json")
    if not isinstance(elonlar,list): elonlar=[]
    elon={"id":len(elonlar)+1,"sarlavha":data.get("sarlavha",""),
          "matn":data.get("matn",""),"ustoz_id":data.get("ustoz_id",0),"vaqt":ts()}
    elonlar.insert(0,elon); save_json("elonlar.json",elonlar)
    threading.Thread(target=_elon_bildirishnoma, args=(elon,), daemon=True).start()
    return {"success":True,"id":elon["id"]}

@app.get("/elon/royxat")
def elon_royxat():
    elonlar=load_json("elonlar.json")
    return {"elonlar":(elonlar if isinstance(elonlar,list) else [])[:20]}

@app.get("/elon/oxirgi_id")
def elon_oxirgi_id():
    elonlar=load_json("elonlar.json")
    if not isinstance(elonlar,list) or not elonlar: return {"id":0}
    return {"id":elonlar[0].get("id",0)}

@app.delete("/elon/ochir/{elon_id}")
def elon_ochir(elon_id: int, ustoz_id: int):
    elonlar=load_json("elonlar.json")
    if isinstance(elonlar,list):
        elonlar=[e for e in elonlar if e.get("id")!=elon_id]
        save_json("elonlar.json",elonlar)
    return {"success":True}

# ── Varaqlar titul ────────────────────────────────────────────────────────────
@app.post("/varaqlar/qosh")
def varaqlar_qosh(data: dict):
    varaqlar=load_json("varaqlar.json")
    soni=str(data.get("soni",""))
    varaqlar[soni]={"fayl_id":data.get("fayl_id",""),"nomi":data.get("nomi","")}
    save_json("varaqlar.json",varaqlar); return {"success":True}

@app.get("/varaqlar/royxat")
def varaqlar_royxat():
    varaqlar=load_json("varaqlar.json")
    natija=[]
    for soni,info in sorted(varaqlar.items(),key=lambda x:int(x[0]) if x[0].isdigit() else 0):
        natija.append({"soni":soni,"fayl_id":info.get("fayl_id",""),"nomi":info.get("nomi","")})
    return {"varaqlar":natija}

@app.delete("/varaqlar/ochir/{soni}")
def varaqlar_ochir(soni: str, ustoz_id: int):
    varaqlar=load_json("varaqlar.json")
    if str(soni) not in varaqlar: raise HTTPException(404,"Varaq topilmadi")
    del varaqlar[str(soni)]
    save_json("varaqlar.json",varaqlar); return {"success":True}

# ── Do'st testi ───────────────────────────────────────────────────────────────
@app.post("/dost/yarat")
def dost_test_yarat(data: DosTTest):
    import string
    kod="DOST_"+''.join(random.choices(string.ascii_uppercase+string.digits,k=6))
    savollar=load_json("savollar.json"); fan_savollar=[]
    if data.fan in savollar:
        for mavzu in savollar[data.fan]: fan_savollar.extend(savollar[data.fan][mavzu])
    tanlangan=random.sample(fan_savollar,min(data.savollar_soni,len(fan_savollar))) if fan_savollar else []
    tests=load_json("tests.json")
    tests[kod]={"fan":data.fan,"tur":"dost","yaratuvchi_id":data.yaratuvchi_id,
                "savollar_soni":len(tanlangan) or data.savollar_soni,"yaratildi":ts(),
                "tugadi":False,"ishtirokchilar":[],"savollar":tanlangan,
                "ustoz_id":data.yaratuvchi_id,"javoblar":{}}
    save_json("tests.json",tests); return {"success":True,"kod":kod}

# ── Trening ───────────────────────────────────────────────────────────────────
@app.get("/trening/{user_id}")
def trening_savollar(user_id: int):
    savollar_baza=load_json("savollar.json"); all_s=[]
    for fan in savollar_baza:
        for mavzu in savollar_baza[fan]: all_s.extend(savollar_baza[fan][mavzu])
    tanlangan=random.sample(all_s,min(20,len(all_s))) if all_s else []
    return {"xato_soni":0,"savollar":tanlangan}

# ── Varaq PDF ─────────────────────────────────────────────────────────────────
@app.get("/varaq/pdf/{kod}")
def varaq_pdf(kod: str):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    tests=load_json("tests.json")
    if kod not in tests: raise HTTPException(404,"Test topilmadi")
    test=tests[kod]; fan=test.get("fan",""); n_savol=test.get("savollar_soni",40); tur=test.get("tur","milliy")
    buf=io.BytesIO(); c=canvas.Canvas(buf,pagesize=A4); W,H=A4
    c.setFont("Helvetica-Bold",18); c.drawCentredString(W/2,H-25*mm,"MILLIY IMTIHON VARAQI")
    c.setFont("Helvetica",12); c.drawCentredString(W/2,H-33*mm,f"Fan: {fan}   |   Test: {kod}")
    for mx,my in [(15*mm,H-15*mm),(W-15*mm,H-15*mm),(15*mm,15*mm),(W-15*mm,15*mm)]:
        c.rect(mx-3*mm,my-3*mm,6*mm,6*mm,fill=1)
    c.setFont("Helvetica-Bold",11); c.drawString(25*mm,H-45*mm,"ID raqami:")
    id_x0=25*mm; id_y0=H-52*mm; rr=3*mm
    for col in range(3):
        cx=id_x0+col*12*mm+5*mm; c.setFont("Helvetica",8)
        for digit in range(10):
            cy=id_y0-digit*7*mm; c.circle(cx,cy,rr,fill=0); c.drawCentredString(cx,cy-2,str(digit))
    ans_y0=H-130*mm; col_w=60*mm; row_h=8*mm
    for s in range(1,n_savol+1):
        col=(s-1)//20; row=(s-1)%20; x0=25*mm+col*col_w; y=ans_y0-row*row_h
        c.setFont("Helvetica-Bold",9); c.drawString(x0,y-2,str(s)+".")
        if tur=="milliy" and 33<=s<=35: harflar=["A","B","C","D","E","F"]
        elif tur=="milliy" and s>=36: c.rect(x0+10*mm,y-3*mm,40*mm,6*mm,fill=0); continue
        else: harflar=["A","B","C","D"]
        for i,h in enumerate(harflar):
            cx=x0+10*mm+i*8*mm; c.circle(cx,y,3*mm,fill=0)
            c.setFont("Helvetica",7); c.drawCentredString(cx,y-2,h)
    c.save(); buf.seek(0)
    return StreamingResponse(buf,media_type="application/pdf",
        headers={"Content-Disposition":f"attachment; filename=varaq_{kod}.pdf"})

@app.post("/ai/savol_yarat")
def ai_savol_yarat(data: dict):
    try:
        from groq import Groq
    except ImportError:
        raise HTTPException(500, "groq kutubxonasi o'rnatilmagan")
    matn = data.get("matn", "").strip()
    fan = data.get("fan", "").strip()
    mavzu = data.get("mavzu", "").strip()
    soni = int(data.get("soni", 5))
    if not matn:
        raise HTTPException(400, "Matn kerak")
    api_key = os.environ.get("GROQ_API_KEY", "")
    if not api_key:
        raise HTTPException(500, "GROQ_API_KEY sozlanmagan")
    try:
        client = Groq(api_key=api_key)
        prompt = f"""Quyidagi matn asosida {soni} ta test savoli tuzing.
Fan: {fan}
Mavzu: {mavzu}
Matn:
{matn[:3000]}
MUHIM: Javobni FAQAT quyidagi JSON formatida bering, boshqa hech narsa yozmang:
[
  {{
    "savol": "Savol matni?",
    "variantlar": {{"A": "birinchi variant", "B": "ikkinchi variant", "C": "uchinchi variant", "D": "to'rtinchi variant"}},
    "togri_javob": "A"
  }}
]"""
        res = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=2000,
        )
        text = res.choices[0].message.content.strip()
        start = text.find("[")
        end = text.rfind("]") + 1
        if start == -1 or end == 0:
            raise HTTPException(500, "AI javob formati noto'g'ri")
        savollar = json.loads(text[start:end])
        return {"success": True, "savollar": savollar}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"AI xato: {str(e)}")
